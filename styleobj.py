#June 2022
#Creates and applies style objects onto a loaded workbook.

import sys
from openpyxl import Workbook,  load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy

wb1 = load_workbook('faithful.xlsx')
ws1 = wb1.active

header_style = NamedStyle(name ="header_style")
header_style.font = Font(bold=True, size=24)
bd = Side(style='thick', color="000000")
header_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
header_style.fill = PatternFill("solid", fgColor="ADD8E6")

body_style = NamedStyle(name = "body_style")
body_style.font = Font(bold=False, size=12)
body_style.fill = PatternFill("solid", fgColor="00FF00FF")

for row in ws1.rows:
    for cell in row:
        cell.style = header_style

wb1.close()
wb1.save('styleobj_out.xlsx') 