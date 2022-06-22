#June 2022
#Testing of merge and alignment functionalities with openpyxl

import sys
from openpyxl import Workbook,  load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy

wb1 = Workbook()
ws1 = wb1.active

ws1.merge_cells('A1:B2')

ws1['A1'] = 3.14
ws1['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws1['A1'].font  = Font(b=True, color="FF0000")

wb1.close()
wb1.save('mergetest_out.xlsx') 