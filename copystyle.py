#June 2022
#copies the style from 'style.xlsx' to 'faithful.xlsx'
import sys
from openpyxl import Workbook,  load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, PatternFill, Border, Side, Alignment, Protection, Font
from copy import copy

wb1 = load_workbook('faithful.xlsx')
ws1 = wb1.active
wb2 = load_workbook('style.xlsx')
ws2 = wb2.active

for row in ws1.rows:
    for cell in row:
        new_cell = ws2.cell(row=cell.row, column=cell.col_idx,
                value= cell.value)
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

wb2.close()
wb2.save('copystyle_out.xlsx') 