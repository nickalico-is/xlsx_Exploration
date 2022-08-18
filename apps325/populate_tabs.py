import shutil
import os
import os.path
from openpyxl import load_workbook

# While testing, if file exists from last run, delete
if os.path.exists('finding_report_v2.xlsx'):
    os.remove('./finding_report_v2.xlsx')

# Create blank copy from template
shutil.copyfile('../finding_report_v2_template.xlsx', './finding_report_v2.xlsx')

intro_tab_data = {
    'site': 'OM Research',
    'recording_equipment_type': 'PC',
    'recording_equipment_number': '123',
    'receiving_data_from_site_since': '01-01-22',
    'last_data_received_date': '02-02-22',
    'report_as_of': '03-03-22',
    'report_sent_through': 'Box',
    'active_trials': 'Galaxi, Quasar'
}

global_overview_data = {
    'site': 'OM Research',
    'total_number_of_patients': 2500,
    'number_of_patients_surfaced': 25,
    'percent_pats_surfaced_of_all_received': 0.01,
    'num_pats_confirmed_eligible_by_coord': 1,
    'percent_confirmed_eligible_by_coord': 0.0004,
}

# Check if file is present
if os.path.exists('finding_report_v2.xlsx'):
    workbook = load_workbook('finding_report_v2.xlsx')
    workbook.active
    print(workbook.sheetnames)
    if "Intro" in workbook.sheetnames:
        intro_tab_worksheet = workbook['Intro']
        intro_tab_worksheet['F7'] = intro_tab_data['site']
        intro_tab_worksheet['F8'] = intro_tab_data['recording_equipment_type']
        intro_tab_worksheet['F9'] = intro_tab_data['recording_equipment_number']
        intro_tab_worksheet['F10'] = intro_tab_data['receiving_data_from_site_since']
        intro_tab_worksheet['F11'] = intro_tab_data['last_data_received_date']
        intro_tab_worksheet['F12'] = intro_tab_data['report_as_of']
        intro_tab_worksheet['F13'] = intro_tab_data['report_sent_through']
        intro_tab_worksheet['F14'] = intro_tab_data['active_trials']
        # for row in intro_tab_worksheet.iter_rows(min_row=7, max_row=14, values_only=True):
        #     print(row)

    if "Global Overview" in workbook.sheetnames:
        global_overview_worksheet = workbook["Global Overview"]
        global_overview_worksheet['B9'] = global_overview_data['total_number_of_patients']
        global_overview_worksheet['C9'] = global_overview_data['number_of_patients_surfaced']
        global_overview_worksheet['D9'] = global_overview_data['percent_pats_surfaced_of_all_received']
        global_overview_worksheet['E9'] = global_overview_data['num_pats_confirmed_eligible_by_coord']
        global_overview_worksheet['F9'] = global_overview_data['percent_confirmed_eligible_by_coord']

    # Hide Percentage columns by default
    if "Global Overview" in workbook.sheetnames:
        for hidden_column in ['D', 'F', 'H', 'J', 'L', 'N', 'P']:
            workbook['Global Overview'].column_dimensions[hidden_column].hidden = True
        
    workbook.close()
    workbook.save('finding_report_v2.xlsx')
else:
    raise Exception("File not present")